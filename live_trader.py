"""
live_trader.py — XGBoost Live Trading Bot
=========================================
Loads trained models and trades live via Pepperstone/cTrader FIX.
Candles are fetched from Yahoo Finance first, with 3 retries (1-min apart),
then fall back to FIX market data when Yahoo is unavailable.

Usage:
    python live_trader.py                        # FIX, both pairs
    python live_trader.py --instrument EUR-USD
    python live_trader.py --dry-run

Setup: copy .env.example to .env and set FIX_PASSWORD (and optional FIX_BALANCE_USD).
Requires outbound SSL to FIX host (default ports 5211 quote, 5212 trade).

Symbols use dash format — EUR-USD maps to FIX symbol id internally.
"""

import os
import sys
import time
import pickle
import logging
import argparse
import warnings
from datetime import datetime, timezone
from typing import Optional

import uuid
import shutil
import threading

import numpy as np
import pandas as pd
from broker.ohlc_utils import normalize_ohlc_df
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── Logging ───────────────────────────────────────────────────────────────────
_LOG_FORMAT  = "%(asctime)s | %(levelname)-8s | %(message)s"
_LOG_DATEFMT = "%Y-%m-%d %H:%M:%S"
logging.basicConfig(level=logging.INFO, format=_LOG_FORMAT, datefmt=_LOG_DATEFMT)
log = logging.getLogger(__name__)

_LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "live_trader.log")
_fh = logging.FileHandler(_LOG_FILE, mode="a", encoding="utf-8")
_fh.setLevel(logging.INFO)
_fh.setFormatter(logging.Formatter(_LOG_FORMAT, datefmt=_LOG_DATEFMT))
logging.getLogger().addHandler(_fh)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
MIN_CONFIDENCE          = 0.65
RISK_PER_TRADE          = 0.005     # 0.5% of account equity per trade
MIN_LOT_SIZE            = 0.01
SELL_TRADES             = True

# Regime gates (must match train_model.py / backtest.py)
REGIME_ADX_MIN          = 18.0
REGIME_ATR_MAX          = 2.0
REGIME_TREND_BIAS       = True
REGIME_EMA_BUY_MIN      = -0.001
REGIME_EMA_SELL_MAX     =  0.001
REGIME_RANGING_ADX30    = 20.0
REGIME_RANGING_EMA_DIST = 0.003
REGIME_EMA_DURATION     = 18

# Monthly profit cap: stop trading a direction once it's up X% this month
CB_MONTHLY_PROFIT_CAP   = 0.08

# Cooldown: bars to skip after a loss
COOLDOWN_BARS           = 3

# Maximum open trades across ALL instruments combined.
# The backtest fires on every signal bar regardless of existing positions,
# which is equivalent to no limit. Set to a low number here to cap real
# exposure, or set to 999 to mirror backtest behaviour exactly.
MAX_CONCURRENT_TRADES   = 999   # set lower (e.g. 3) to cap live exposure

# Lookback candles for indicators
CANDLE_LOOKBACK         = 600

# How often to check for a new completed candle (seconds)
POLL_INTERVAL_SECONDS   = 30

# Instrument map: model name → MT5 symbol name
INSTRUMENT_MAP = {
    "EUR-USD": "EURUSD",
    "GBP-USD": "GBPUSD",
}

BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
MODEL_CACHE_DIR = os.path.join(BASE_DIR, "models")
FIX_CACHE_DIR   = os.path.join(BASE_DIR, "data", "fix_cache")

TRADE_LOG_XLSX  = os.path.join(BASE_DIR, "trade_log.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL TRADE LOGGER
# ─────────────────────────────────────────────────────────────────────────────

XLSX_COLUMNS = [
    "Timestamp (UTC)", "Symbol", "Direction", "Lot Size",
    "Entry Price", "SL Price", "TP Price",
    "Stop (pips)", "TP (pips)", "ATR",
    "Confidence", "Buy Prob", "Sell Prob",
    "ADX", "ATR Regime", "EMA 200 Dist",
    "Account Equity", "Risk Amount", "Currency",
    "Dry Run", "Signal Reason",
]

_HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
_ALT_FILL      = PatternFill("solid", fgColor="EBF0F8")   # light blue-grey
_HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_BODY_FONT     = Font(name="Arial", size=10)
_BUY_FONT      = Font(name="Arial", size=10, color="1A7F3C", bold=True)
_SELL_FONT     = Font(name="Arial", size=10, color="C0392B", bold=True)
_THIN          = Side(border_style="thin", color="D0D0D0")
_BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

COL_WIDTHS = {
    "Timestamp (UTC)": 22, "Symbol": 9, "Direction": 10, "Lot Size": 9,
    "Entry Price": 12, "SL Price": 12, "TP Price": 12,
    "Stop (pips)": 11, "TP (pips)": 11, "ATR": 10,
    "Confidence": 11, "Buy Prob": 10, "Sell Prob": 10,
    "ADX": 8, "ATR Regime": 11, "EMA 200 Dist": 13,
    "Account Equity": 15, "Risk Amount": 12, "Currency": 9,
    "Dry Run": 9, "Signal Reason": 40,
}


def _apply_header(ws) -> None:
    """Apply styled header row to the worksheet."""
    for col_idx, col_name in enumerate(XLSX_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 14)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _style_data_row(ws, row_idx: int, direction: str) -> None:
    """Apply alternating background and direction-colour to a data row."""
    fill = _ALT_FILL if row_idx % 2 == 0 else None
    for col_idx in range(1, len(XLSX_COLUMNS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border    = _BORDER
        cell.alignment = Alignment(vertical="center")
        if fill:
            cell.fill = fill
        # Direction column gets colour
        if col_idx == 3:
            cell.font = _BUY_FONT if direction == "BUY" else _SELL_FONT
        else:
            cell.font = _BODY_FONT


def _is_file_locked(path: str) -> bool:
    """Try to open the file exclusively; return True if it's locked."""
    if not os.path.exists(path):
        return False
    try:
        with open(path, "r+b"):
            return False
    except (IOError, PermissionError):
        return True


class ExcelTradeLogger:
    """
    Thread-safe Excel trade logger with locked-file fallback.

    Behaviour when trade_log.xlsx is open (locked by Excel):
      • Writes to a random-named temp file instead.
      • Spawns a background thread to keep retrying the merge until the
        main file becomes writable.
      • If another trade fires while a temp is still pending, the old temp
        is deleted and a fresh one is created (so only one pending temp
        exists at any time).
    """

    def __init__(self, target_path: str):
        self.target_path   = target_path
        self._lock         = threading.Lock()   # guards pending_temp state
        self._pending_temp: Optional[str] = None
        self._merge_thread: Optional[threading.Thread] = None

    # ── public ────────────────────────────────────────────────────────

    def log_trade(self, row_data: dict) -> None:
        """Append one trade row. Call from any thread."""
        with self._lock:
            if _is_file_locked(self.target_path):
                self._handle_locked(row_data)
            else:
                self._write_or_append(self.target_path, row_data)
                log.info(f"[TradeLog] Written to {os.path.basename(self.target_path)}")

    # ── private ───────────────────────────────────────────────────────

    def _handle_locked(self, row_data: dict) -> None:
        """Target is locked — write to a fresh temp and start/restart merge watcher."""
        # Discard any stale pending temp
        if self._pending_temp and os.path.exists(self._pending_temp):
            try:
                os.remove(self._pending_temp)
            except OSError:
                pass
            log.info(f"[TradeLog] Dropped stale temp {os.path.basename(self._pending_temp)}")

        # Create new temp
        temp_name = f"trade_log_tmp_{uuid.uuid4().hex[:8]}.xlsx"
        temp_path = os.path.join(os.path.dirname(self.target_path), temp_name)
        self._write_or_append(temp_path, row_data)
        self._pending_temp = temp_path
        log.warning(
            f"[TradeLog] trade_log.xlsx is open — buffered to {temp_name}"
        )

        # Start (or restart) background merge watcher
        if self._merge_thread is None or not self._merge_thread.is_alive():
            self._merge_thread = threading.Thread(
                target=self._merge_watcher, daemon=True
            )
            self._merge_thread.start()

    def _merge_watcher(self) -> None:
        """Background: keep trying to merge the pending temp into the main file."""
        while True:
            time.sleep(15)   # check every 15 seconds
            with self._lock:
                if self._pending_temp is None:
                    return
                if not os.path.exists(self._pending_temp):
                    self._pending_temp = None
                    return
                if _is_file_locked(self.target_path):
                    log.debug("[TradeLog] Still locked, will retry...")
                    continue
                # Merge temp rows into main file
                try:
                    self._merge_temp_into_main(self._pending_temp)
                    log.info(
                        f"[TradeLog] Merged {os.path.basename(self._pending_temp)} "
                        f"→ trade_log.xlsx"
                    )
                    os.remove(self._pending_temp)
                    self._pending_temp = None
                    return
                except Exception as exc:
                    log.error(f"[TradeLog] Merge failed: {exc}")

    def _merge_temp_into_main(self, temp_path: str) -> None:
        """Read all data rows from temp and append them to the main xlsx."""
        wb_tmp = openpyxl.load_workbook(temp_path)
        ws_tmp = wb_tmp.active
        # Rows 2+ are data rows (row 1 is header)
        data_rows = list(ws_tmp.iter_rows(min_row=2, values_only=True))
        wb_tmp.close()

        if not data_rows:
            return

        if not os.path.exists(self.target_path):
            # Nothing to merge into; just rename the temp
            shutil.move(temp_path, self.target_path)
            return

        wb = openpyxl.load_workbook(self.target_path)
        ws = wb.active
        next_row = ws.max_row + 1
        for row_values in data_rows:
            direction = row_values[2] if len(row_values) > 2 else "BUY"
            for col_idx, val in enumerate(row_values, start=1):
                ws.cell(row=next_row, column=col_idx, value=val)
            _style_data_row(ws, next_row, direction)
            next_row += 1
        wb.save(self.target_path)
        wb.close()

    @staticmethod
    def _write_or_append(path: str, row_data: dict) -> None:
        """Write row_data to path, creating the workbook with a header if needed."""
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            row_idx = ws.max_row + 1
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Trades"
            _apply_header(ws)
            row_idx = 2

        direction = row_data.get("Direction", "BUY")
        for col_idx, col_name in enumerate(XLSX_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
        _style_data_row(ws, row_idx, direction)

        wb.save(path)
        wb.close()


# Singleton loggers used throughout the module
_excel_logger = ExcelTradeLogger(TRADE_LOG_XLSX)
_firestore_logger = None
try:
    from integrations.trade_log import CompositeTradeLogger
    from integrations.firestore_logger import FirestoreLogger

    _firestore_logger = FirestoreLogger()
    _trade_logger = CompositeTradeLogger(_excel_logger, _firestore_logger)
except ImportError:
    _trade_logger = _excel_logger


# ─────────────────────────────────────────────────────────────────────────────
# FEATURE ENGINEERING  (exact copy from train_model.py / backtest.py)
# ─────────────────────────────────────────────────────────────────────────────

def _rsi(series, period):
    delta = series.diff()
    gain  = delta.clip(lower=0).ewm(com=period - 1, adjust=False).mean()
    loss  = (-delta.clip(upper=0)).ewm(com=period - 1, adjust=False).mean()
    return 100 - (100 / (1 + gain / (loss + 1e-9)))


def _atr(h, l, c, period):
    tr = pd.concat([h - l, (h - c.shift()).abs(), (l - c.shift()).abs()], axis=1).max(axis=1)
    return tr.ewm(com=period - 1, adjust=False).mean()


def _bb_width(series, period):
    sma = series.rolling(period).mean()
    std = series.rolling(period).std()
    return (2 * std) / (sma + 1e-9)


def _adx(h, l, c, period=14):
    up   = h.diff()
    down = -l.diff()
    plus_dm  = np.where((up > down) & (up > 0), up, 0.0)
    minus_dm = np.where((down > up) & (down > 0), down, 0.0)
    tr = pd.concat([h - l, (h - c.shift()).abs(), (l - c.shift()).abs()], axis=1).max(axis=1)
    atr_s    = pd.Series(tr.values).ewm(com=period - 1, adjust=False).mean()
    plus_di  = 100 * pd.Series(plus_dm).ewm(com=period - 1, adjust=False).mean() / (atr_s + 1e-9)
    minus_di = 100 * pd.Series(minus_dm).ewm(com=period - 1, adjust=False).mean() / (atr_s + 1e-9)
    dx  = 100 * (plus_di - minus_di).abs() / (plus_di + minus_di + 1e-9)
    adx = dx.ewm(com=period - 1, adjust=False).mean()
    adx.index = c.index
    return adx


def add_features(df: pd.DataFrame, pip_size: float) -> pd.DataFrame:
    """Build all features. df must have: open, high, low, close, volume, ask_close, ask_high, ask_low, spread_pips"""
    df = normalize_ohlc_df(df, pip_size)
    c, h, l, o, v = df["close"], df["high"], df["low"], df["open"], df["volume"]

    df["ema_9"]      = c.ewm(span=9).mean()
    df["ema_21"]     = c.ewm(span=21).mean()
    df["ema_cross"]  = (df["ema_9"] - df["ema_21"]) / c
    df["rsi_14"]     = _rsi(c, 14)
    df["rsi_slope"]  = df["rsi_14"].diff(3)

    macd_line        = c.ewm(span=12).mean() - c.ewm(span=26).mean()
    signal_line      = macd_line.ewm(span=9).mean()
    df["macd_hist"]  = macd_line - signal_line
    df["macd_slope"] = df["macd_hist"].diff(2)

    df["roc_5"]      = c.pct_change(5)
    df["roc_10"]     = c.pct_change(10)
    df["atr_14"]     = _atr(h, l, c, 14)
    df["atr_norm"]   = df["atr_14"] / c
    df["bb_width"]   = _bb_width(c, 20)
    df["high_low_r"] = (h - l) / c
    # Volume from free FX feeds is often 0; use min 1.0 so vol_ratio is defined.
    vol_ma = v.rolling(20, min_periods=1).mean().clip(lower=1.0)
    df["vol_ratio"]  = v / vol_ma
    df["vol_slope"]  = v.pct_change(3).replace([np.inf, -np.inf], np.nan).fillna(0.0)

    df["hour"]       = df.index.hour
    df["is_london"]  = df["hour"].between(7, 16).astype(int)
    df["is_ny"]      = df["hour"].between(13, 21).astype(int)
    df["is_overlap"] = df["hour"].between(13, 16).astype(int)

    for lag in [1, 2, 3, 5]:
        df[f"ret_lag{lag}"] = c.pct_change(lag)

    df["body_ratio"] = (c - o).abs() / (h - l + 1e-9)
    df["upper_wick"] = (h - df[["close", "open"]].max(axis=1)) / (h - l + 1e-9)
    df["lower_wick"] = (df[["close", "open"]].min(axis=1) - l) / (h - l + 1e-9)

    df["spread_norm"]  = df["spread_pips"] * pip_size / c
    df["spread_ma"]    = df["spread_pips"].rolling(20).mean()
    df["spread_ratio"] = df["spread_pips"] / (df["spread_ma"] + 1e-9)

    ema_200              = c.ewm(span=200).mean()
    df["ema_200_dist"]   = (c - ema_200) / c
    df["ema_200_slope"]  = ema_200.pct_change(10)
    df["adx_14"]         = _adx(h, l, c, 14)
    df["roc_48"]         = c.pct_change(48)
    df["roc_120"]        = c.pct_change(120)

    atr_50_ma            = df["atr_14"].rolling(50).mean()
    df["atr_regime"]     = df["atr_14"] / (atr_50_ma + 1e-9)
    df["adx_30ma"]       = df["adx_14"].rolling(30).mean()
    df["ema_trend_consistency"] = df["ema_200_slope"].rolling(20).apply(
        lambda x: float((x > 0).sum()) / max(len(x), 1), raw=True
    )
    return df


from broker.factory import get_broker, init_broker, load_dotenv
from broker.history import fetch_hourly_history

# ─────────────────────────────────────────────────────────────────────────────
# CANDLE FETCH WITH RETRY + FALLBACK
# ─────────────────────────────────────────────────────────────────────────────

def fetch_candles_with_fallback(symbol: str, count: int = 600, max_retries: int = 3) -> Optional[pd.DataFrame]:
    """Try Yahoo Finance first, retry up to `max_retries` times with 1-min delay, then fall back to FIX broker."""
    for attempt in range(1, max_retries + 1):
        df = fetch_hourly_history(symbol, count)
        if df is not None and len(df) >= count:
            return df
        if attempt < max_retries:
            log.warning(
                "%s | Yahoo Finance failed (attempt %d/%d) — retrying in 60s...",
                symbol, attempt, max_retries,
            )
            time.sleep(60)

    log.warning("%s | Yahoo Finance exhausted — falling back to FIX broker candles", symbol)
    return get_broker().get_bid_ask_candles(symbol, count)


# ─────────────────────────────────────────────────────────────────────────────
# CANDLE PROGRESS BAR
# ───────────────────────────────────────────────────────────────────────────────

def _draw_candle_bar(symbol: str, candle_open_time, open_trades: int) -> None:
    """
    Overwrite the current terminal line with a filling progress bar that shows
    how far through the current 1H candle we are.

    Uses \r (carriage return) to stay on one line — only goes to stdout,
    never to the log file, so the log stays clean.

    Example output:
      EURUSD [===========-------] 58% | 25min left | trades=1
    """
    now        = datetime.now(timezone.utc)
    # candle_open_time is the timestamp of the current (forming) candle
    elapsed    = (now - candle_open_time).total_seconds()
    total      = 3600.0   # 1H candle = 3600 seconds
    pct        = min(elapsed / total, 1.0)
    mins_left  = max(0, int((total - elapsed) / 60))

    width      = 20   # bar fill width in characters
    filled     = int(pct * width)
    bar        = "█" * filled + "░" * (width - filled)

    line = (
        f"\r  {symbol} [[32m{bar}[0m] {pct*100:4.0f}% | "
        f"{mins_left:2d}min left | trades={open_trades}   "
    )
    sys.stdout.write(line)
    sys.stdout.flush()


# SIGNAL GENERATION
# ─────────────────────────────────────────────────────────────────────────────

def load_model(instrument_key: str) -> Optional[dict]:
    """Load live_models.pkl for an instrument. Returns payload dict or None."""
    path = os.path.join(MODEL_CACHE_DIR, instrument_key, "live_models.pkl")
    if not os.path.isfile(path):
        log.error(f"Model not found: {path}")
        log.error("Run train_model.py first.")
        return None
    with open(path, "rb") as fh:
        payload = pickle.load(fh)
    log.info(f"Model loaded: {path}")
    return payload


def get_signal(df: pd.DataFrame, model_payload: dict) -> dict:
    """
    Compute feature vector from the LAST completed candle and run the models.

    Returns dict with keys:
        direction    : "BUY", "SELL", or "HOLD"
        confidence   : float (max of buy_prob / sell_prob)
        buy_prob     : float
        sell_prob    : float
        atr          : float (ATR value for position sizing)
        regime_ok    : bool (True if all regime filters passed)
        regime_reason: str  (reason for rejection if regime_ok=False)
        ema_200_dist : float
        adx          : float
        atr_regime   : float
        adx_30ma     : float
    """
    pip_size      = model_payload.get("pip_size", 0.0001)
    contract_size = model_payload.get("contract_size", 100_000)
    quote_is_usd  = model_payload.get("quote_is_usd", True)
    feature_cols  = model_payload["feature_cols"]
    buy_model     = model_payload["buy_model"]
    sell_model    = model_payload["sell_model"]
    scaler        = model_payload["scaler"]

    # Regime thresholds — read from payload so they always match backtest.
    # Falls back to the module-level defaults for old .pkl files.
    reg_adx_min      = model_payload.get("REGIME_ADX_MIN",          REGIME_ADX_MIN)
    reg_atr_max      = model_payload.get("REGIME_ATR_MAX",          REGIME_ATR_MAX)
    reg_trend_bias   = model_payload.get("REGIME_TREND_BIAS",       REGIME_TREND_BIAS)
    reg_ema_buy_min  = model_payload.get("REGIME_EMA_BUY_MIN",      REGIME_EMA_BUY_MIN)
    reg_ema_sell_max = model_payload.get("REGIME_EMA_SELL_MAX",     REGIME_EMA_SELL_MAX)
    reg_ranging_adx  = model_payload.get("REGIME_RANGING_ADX30",    REGIME_RANGING_ADX30)
    reg_ranging_ema  = model_payload.get("REGIME_RANGING_EMA_DIST", REGIME_RANGING_EMA_DIST)
    reg_ema_duration = model_payload.get("REGIME_EMA_DURATION",     REGIME_EMA_DURATION)

    # SL/TP multipliers
    atr_sl_mult = model_payload.get("atr_sl_mult", 1.5)
    tp_rr       = model_payload.get("tp_rr",       1.5)

    # Build features on the full history, then take the last row
    df_norm = normalize_ohlc_df(df.copy(), pip_size)
    df_feat = add_features(df_norm, pip_size)
    before_drop = len(df_feat)
    df_feat = df_feat.dropna()
    if len(df_feat) == 0:
        log.warning(
            "Feature rows empty after dropna (input bars=%d, rows before dropna=%d). "
            "Delete data/fix_cache/*.pkl and restart if this persists.",
            len(df_norm),
            before_drop,
        )
        return _empty_signal("Not enough data for features")

    row = df_feat.iloc[-1]  # last completed candle

    # Check all required features exist
    missing = [f for f in feature_cols if f not in df_feat.columns]
    if missing:
        log.warning(f"Missing features: {missing}")
        return _empty_signal(f"Missing features: {missing}")

    X = scaler.transform(row[feature_cols].values.reshape(1, -1))

    buy_prob  = float(buy_model.predict_proba(X)[0, 1])
    sell_prob = float(sell_model.predict_proba(X)[0, 1])

    # Extract regime indicators from last row
    adx        = float(row.get("adx_14", 25.0))
    atr_regime = float(row.get("atr_regime", 1.0))
    ema_dist   = float(row.get("ema_200_dist", 0.0))
    adx_30ma   = float(row.get("adx_30ma", 25.0))
    atr        = float(row.get("atr_14", 0.0008))

    # ── Determine raw signal ──────────────────────────────────────────
    buy_fires  = buy_prob  >= MIN_CONFIDENCE
    sell_fires = sell_prob >= MIN_CONFIDENCE and SELL_TRADES

    if buy_fires and sell_fires:
        raw_direction = "BUY" if buy_prob >= sell_prob else "SELL"
        confidence    = max(buy_prob, sell_prob)
    elif buy_fires:
        raw_direction = "BUY"
        confidence    = buy_prob
    elif sell_fires:
        raw_direction = "SELL"
        confidence    = sell_prob
    else:
        return {
            "direction": "HOLD", "confidence": max(buy_prob, sell_prob),
            "buy_prob": buy_prob, "sell_prob": sell_prob,
            "regime_ok": True, "regime_reason": "Below confidence threshold",
            "atr": atr, "ema_200_dist": ema_dist, "adx": adx,
            "atr_regime": atr_regime, "adx_30ma": adx_30ma,
        }

    dir_int = 1 if raw_direction == "BUY" else -1

    # ── Regime gate (same 5 layers as backtest.py) ─────────────────────
    if adx < reg_adx_min:
        return _hold_result(buy_prob, sell_prob, atr, ema_dist, adx, atr_regime, adx_30ma,
                            f"ADX {adx:.1f} < {reg_adx_min} (ranging market)")

    if atr_regime > reg_atr_max:
        return _hold_result(buy_prob, sell_prob, atr, ema_dist, adx, atr_regime, adx_30ma,
                            f"ATR regime {atr_regime:.2f} > {reg_atr_max} (volatility spike)")

    if adx_30ma < reg_ranging_adx and abs(ema_dist) < reg_ranging_ema:
        return _hold_result(buy_prob, sell_prob, atr, ema_dist, adx, atr_regime, adx_30ma,
                            f"Sustained ranging (adx_30ma={adx_30ma:.1f}, ema_dist={ema_dist:.4f})")

    if reg_trend_bias:
        if dir_int == 1 and ema_dist < reg_ema_buy_min:
            return _hold_result(buy_prob, sell_prob, atr, ema_dist, adx, atr_regime, adx_30ma,
                                f"BUY blocked: ema_dist {ema_dist:.4f} < {reg_ema_buy_min}")
        if dir_int == -1 and ema_dist > reg_ema_sell_max:
            return _hold_result(buy_prob, sell_prob, atr, ema_dist, adx, atr_regime, adx_30ma,
                                f"SELL blocked: ema_dist {ema_dist:.4f} > {reg_ema_sell_max}")

        # EMA slope filter (use last 48 rows of df_feat)
        if len(df_feat) >= 48:
            ema_dist_arr = df_feat["ema_200_dist"].values
            ema_slope = ema_dist_arr[-1] - ema_dist_arr[-49]
            if dir_int == 1 and ema_slope < -0.003:
                return _hold_result(buy_prob, sell_prob, atr, ema_dist, adx, atr_regime, adx_30ma,
                                    f"BUY blocked: EMA slope accelerating down ({ema_slope:.4f})")
            if dir_int == -1 and ema_slope > 0.003:
                return _hold_result(buy_prob, sell_prob, atr, ema_dist, adx, atr_regime, adx_30ma,
                                    f"SELL blocked: EMA slope accelerating up ({ema_slope:.4f})")

    # EMA duration check (last 24 bars)
    if len(df_feat) >= 24:
        lookback = df_feat["ema_200_dist"].values[-24:]
        if dir_int == 1 and int((lookback < 0).sum()) > reg_ema_duration:
            return _hold_result(buy_prob, sell_prob, atr, ema_dist, adx, atr_regime, adx_30ma,
                                f"BUY blocked: price below EMA-200 for >{reg_ema_duration}/24 bars")
        if dir_int == -1 and int((lookback > 0).sum()) > reg_ema_duration:
            return _hold_result(buy_prob, sell_prob, atr, ema_dist, adx, atr_regime, adx_30ma,
                                f"SELL blocked: price above EMA-200 for >{reg_ema_duration}/24 bars")

    return {
        "direction":    raw_direction,
        "confidence":   confidence,
        "buy_prob":     buy_prob,
        "sell_prob":    sell_prob,
        "regime_ok":    True,
        "regime_reason": "All filters passed",
        "atr":          atr,
        "ema_200_dist": ema_dist,
        "adx":          adx,
        "atr_regime":   atr_regime,
        "adx_30ma":     adx_30ma,
    }


def _hold_result(buy_prob, sell_prob, atr, ema_dist, adx, atr_regime, adx_30ma, reason):
    return {
        "direction": "HOLD", "confidence": max(buy_prob, sell_prob),
        "buy_prob": buy_prob, "sell_prob": sell_prob,
        "regime_ok": False, "regime_reason": reason,
        "atr": atr, "ema_200_dist": ema_dist, "adx": adx,
        "atr_regime": atr_regime, "adx_30ma": adx_30ma,
    }


def _empty_signal(reason: str) -> dict:
    """Signal dict when features/models cannot run (avoids KeyError downstream)."""
    return {
        "direction": "HOLD",
        "confidence": 0.0,
        "buy_prob": 0.0,
        "sell_prob": 0.0,
        "regime_ok": False,
        "regime_reason": reason,
        "atr": 0.0,
        "ema_200_dist": 0.0,
        "adx": 0.0,
        "atr_regime": 0.0,
        "adx_30ma": 0.0,
    }


# ─────────────────────────────────────────────────────────────────────────────
# POSITION SIZING
# ─────────────────────────────────────────────────────────────────────────────

def compute_lot_size(atr: float, pip_size: float, balance: float,
                     contract_size: int = 100_000,
                     atr_sl_mult: float = 1.5,
                     tp_rr: float = 1.5) -> tuple:
    """
    Compute lot size and SL/TP distances using values from the model payload.
    """
    atr        = max(atr, 5 * pip_size)
    stop_dist  = atr_sl_mult * atr
    tp_dist    = tp_rr * stop_dist
    stop_pips  = stop_dist / pip_size
    tp_pips    = tp_dist   / pip_size

    risk_amount = balance * RISK_PER_TRADE   # in account currency (USD or other)
    # contract_size and multipliers come from the model payload (set at train time)
    lot_size = risk_amount / (stop_dist * contract_size)
    lot_size = max(lot_size, MIN_LOT_SIZE)

    return lot_size, stop_dist, tp_dist, stop_pips, tp_pips


# ─────────────────────────────────────────────────────────────────────────────
# MAIN TRADING LOOP
# ─────────────────────────────────────────────────────────────────────────────

class InstrumentState:
    """Tracks per-instrument state: cooldown and monthly profit cap."""
    def __init__(self, instrument_key: str, broker_symbol: str):
        self.instrument_key = instrument_key
        self.broker_symbol  = broker_symbol
        self.last_candle_time = None
        self.last_loss_bar    = -999
        self.bar_index        = 0

        # Monthly profit cap
        self.current_month   = None
        self.buy_month_pnl   = 0.0
        self.sell_month_pnl  = 0.0
        self.buy_month_cap   = 0.0
        self.sell_month_cap  = 0.0
        self.last_wait_logged   = None  # candle time we last logged a waiting line
        self.last_heartbeat_time = None  # datetime of last heartbeat log


def run_instrument_loop(state: InstrumentState, model_payload: dict,
                        dry_run: bool = False) -> None:
    """Check for a new completed candle and act if a signal fires."""
    broker = get_broker()
    symbol = state.broker_symbol
    now_utc = datetime.now(timezone.utc)

    # ── Lightweight time-based check (zero I/O) ────────────────────────
    # Avoids downloading 650 Yahoo candles on every poll cycle. Since 1H
    # candles close at the top of each hour, we can compute when the next
    # one is due without any network calls.
    if state.last_candle_time is not None:
        next_candle_close = state.last_candle_time + pd.Timedelta(hours=1)
        if now_utc < next_candle_close:
            if state.last_wait_logged != state.last_candle_time:
                state.last_wait_logged = state.last_candle_time
                state.last_heartbeat_time = now_utc
                mins_to_next = int((next_candle_close - now_utc).total_seconds() / 60)
                log.info(
                    f"{symbol} | Waiting for next candle "
                    f"(last closed: {state.last_candle_time}) "
                    f"~{mins_to_next}min remaining"
                )
            return

        mins_since_last = (now_utc - state.last_candle_time).total_seconds() / 60
        if mins_since_last > 120:
            if state.last_heartbeat_time is None or \
                    (now_utc - state.last_heartbeat_time).total_seconds() >= 3600:
                state.last_heartbeat_time = now_utc
                log.warning(
                    f"{symbol} | Market appears closed or price feed stopped. "
                    f"Last candle: {state.last_candle_time} ({mins_since_last/60:.1f}h ago). "
                    f"Bot is paused — will resume when market reopens."
                )
            return

    # ── Full Yahoo fetch (once per new candle or on first run) ──────────
    df = fetch_candles_with_fallback(symbol, CANDLE_LOOKBACK)
    if df is None or len(df) < 300:
        log.warning(f"{symbol}: Not enough candles ({len(df) if df is not None else 0})")
        return

    latest_closed_time = df.index[-2]

    if state.last_candle_time is not None and state.last_candle_time == latest_closed_time:
        return  # Somehow the same candle — skip

    state.last_candle_time = latest_closed_time
    state.bar_index += 1
    log.info(f"\n{'─'*55}")
    log.info(f"{symbol} | 1H candle closed: {latest_closed_time}")

    # Use all data up to and including the last closed candle
    df_closed = df.iloc[:-1]

    # Get signal from the model
    signal = get_signal(df_closed, model_payload)

    pip_size = model_payload.get("pip_size", 0.0001)
    log.info(
        f"{symbol} | BUY={signal.get('buy_prob', 0):.3f} SELL={signal.get('sell_prob', 0):.3f} "
        f"→ {signal['direction']} (conf={signal.get('confidence', 0):.3f})"
    )
    log.info(
        f"  ADX={signal.get('adx', 0):.1f}  ATR_reg={signal.get('atr_regime', 0):.2f}  "
        f"EMA_dist={signal.get('ema_200_dist', 0):.4f}  ADX_30ma={signal.get('adx_30ma', 0):.1f}"
    )

    if signal["direction"] == "HOLD":
        log.info(f"  → HOLD: {signal['regime_reason']}")
        return

    direction = signal["direction"]

    # ── Concurrent position gate ─────────────────────────────────────────
    # Count all open positions across every instrument this bot manages.
    # This mirrors the backtest which allows unlimited simultaneous trades
    # (MAX_CONCURRENT_TRADES = 999 by default). Lower it to cap live exposure.
    broker.refresh_positions()
    bot_positions = broker.get_bot_positions()
    if len(bot_positions) >= MAX_CONCURRENT_TRADES:
        log.info(
            f"  → HOLD: {len(bot_positions)} open position(s) already "
            f"(MAX_CONCURRENT_TRADES={MAX_CONCURRENT_TRADES})"
        )
        return

    # ── Cooldown check ────────────────────────────────────────────────
    bars_since_loss = state.bar_index - state.last_loss_bar
    if bars_since_loss < COOLDOWN_BARS:
        log.info(f"  → HOLD: cooldown ({bars_since_loss}/{COOLDOWN_BARS} bars since last loss)")
        return

    # ── Monthly profit cap ────────────────────────────────────────────
    now = datetime.now(timezone.utc)
    if state.current_month != (now.year, now.month):
        state.current_month  = (now.year, now.month)
        state.buy_month_pnl  = 0.0
        state.sell_month_pnl = 0.0
        balance = broker.get_account_balance()
        state.buy_month_cap  = balance
        state.sell_month_cap = balance

    if CB_MONTHLY_PROFIT_CAP is not None:
        if direction == "BUY" and state.buy_month_cap > 0:
            if state.buy_month_pnl >= CB_MONTHLY_PROFIT_CAP * state.buy_month_cap:
                log.info(f"  → HOLD: BUY monthly profit cap hit ({state.buy_month_pnl:.2f})")
                return
        if direction == "SELL" and state.sell_month_cap > 0:
            if state.sell_month_pnl >= CB_MONTHLY_PROFIT_CAP * state.sell_month_cap:
                log.info(f"  → HOLD: SELL monthly profit cap hit ({state.sell_month_pnl:.2f})")
                return

    # ── Position sizing ───────────────────────────────────────────────
    balance  = broker.get_account_balance()
    currency = broker.get_account_currency()
    tick     = broker.get_tick(symbol)
    if tick is None:
        log.error(f"  → HOLD: no live quote for {symbol}")
        return
    price    = tick.ask if direction == "BUY" else tick.bid

    lot_size, stop_dist, tp_dist, stop_pips, tp_pips = compute_lot_size(
        atr           = signal["atr"],
        pip_size      = model_payload.get("pip_size", 0.0001),
        balance       = balance,
        contract_size = model_payload.get("contract_size", 100_000),
        atr_sl_mult   = model_payload.get("atr_sl_mult", 1.5),
        tp_rr         = model_payload.get("tp_rr", 1.5),
    )

    # SL/TP prices
    if direction == "BUY":
        sl_price = price - stop_dist
        tp_price = price + tp_dist
    else:
        sl_price = price + stop_dist
        tp_price = price - tp_dist

    log.info(
        f"  → TRADE: {direction} {lot_size:.2f} lots {symbol} @ ~{price:.5f} "
        f"| SL={sl_price:.5f} ({stop_pips:.1f}p) | TP={tp_price:.5f} ({tp_pips:.1f}p) "
        f"| Risk={balance * RISK_PER_TRADE:.2f} {currency}"
    )

    # ── Place order ───────────────────────────────────────────────────
    success = broker.place_order(
        symbol=symbol, direction=direction,
        lot_size=lot_size, sl_price=sl_price, tp_price=tp_price,
        comment=f"XGB_{state.instrument_key}", dry_run=dry_run
    )

    if not success:
        log.error(f"  Order failed for {symbol}")
        return

    # ── Pull deal history from Open API (live only) ───────────────────
    if not dry_run:
        try:
            deals = broker.get_recent_deals(hours=1, max_rows=5)
            if deals:
                d = deals[0]
                log.info(
                    "  Deal confirmed: id=%s %s %.2f lots @ %.5f",
                    d.get("deal_id", "?"), d.get("direction", "?"),
                    d.get("volume", 0), d.get("price", 0),
                )
            else:
                log.info("  No deals yet — may appear next poll cycle")
        except Exception as exc:
            log.debug("Deal history pull failed: %s", exc)

    # ── Log trade to Excel ────────────────────────────────────────────
    risk_amount = balance * RISK_PER_TRADE
    row = {
        "Timestamp (UTC)":   datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
        "Symbol":            symbol,
        "Direction":         direction,
        "Lot Size":          round(lot_size, 2),
        "Entry Price":       round(price, 5),
        "SL Price":          round(sl_price, 5),
        "TP Price":          round(tp_price, 5),
        "Stop (pips)":       round(stop_pips, 1),
        "TP (pips)":         round(tp_pips, 1),
        "ATR":               round(signal["atr"], 6),
        "Confidence":        round(signal["confidence"], 4),
        "Buy Prob":          round(signal.get("buy_prob", 0.0), 4),
        "Sell Prob":         round(signal.get("sell_prob", 0.0), 4),
        "ADX":               round(signal["adx"], 2),
        "ATR Regime":        round(signal["atr_regime"], 4),
        "EMA 200 Dist":      round(signal["ema_200_dist"], 6),
        "Account Equity":    round(balance, 2),
        "Risk Amount":       round(risk_amount, 2),
        "Currency":          currency,
        "Dry Run":           "Yes" if dry_run else "No",
        "Signal Reason":     signal.get("regime_reason", "All filters passed"),
    }
    _trade_logger.log_trade(row)


def main():
    parser = argparse.ArgumentParser(description="XGBoost Live Trader (cTrader FIX)")
    parser.add_argument("--instrument", default=None,
                        help="Single instrument key (e.g. EUR-USD). Default: all.")
    parser.add_argument("--dry-run", action="store_true", default=False,
                        help="Print signals without placing orders.")
    args = parser.parse_args()

    load_dotenv(os.path.join(BASE_DIR, ".env"))

    # Stream logs to Firestore daily buckets (UTC) when Firebase is enabled
    if _firestore_logger is not None and _firestore_logger.enabled:
        try:
            from integrations.firestore_log_handler import attach_firestore_daily_logs
            attach_firestore_daily_logs(_firestore_logger)
        except Exception as exc:
            log.warning("Firestore daily log handler not attached: %s", exc)

    # Determine instruments to trade
    if args.instrument:
        if args.instrument not in INSTRUMENT_MAP:
            # Try reverse lookup (user may pass MT5 symbol like EURUSD)
            rev = {v: k for k, v in INSTRUMENT_MAP.items()}
            if args.instrument in rev:
                instruments = {rev[args.instrument]: INSTRUMENT_MAP[rev[args.instrument]]}
            else:
                log.error(f"Unknown instrument: {args.instrument}")
                log.error(f"Valid keys: {list(INSTRUMENT_MAP.keys())}")
                sys.exit(1)
        else:
            instruments = {args.instrument: INSTRUMENT_MAP[args.instrument]}
    else:
        instruments = INSTRUMENT_MAP

    broker_symbols = list(instruments.values())
    broker = init_broker(broker_symbols, FIX_CACHE_DIR)
    if not broker.connect():
        log.error("Could not connect to broker (FIX). Check your .env file and network connection.")
        sys.exit(1)

    log.info(f"Broker: {broker.name}")
    log.info(f"Trading instruments: {list(instruments.keys())}")
    if args.dry_run:
        log.info("DRY RUN mode — no orders will be placed.")

    # Load models
    models = {}
    for key in instruments:
        payload = load_model(key)
        if payload is None:
            log.error(f"Cannot load model for {key}. Run train_model.py first.")
            continue
        models[key] = payload

    if not models:
        log.error("No models loaded. Exiting.")
        sys.exit(1)

    # Initialise per-instrument state
    states = {
        key: InstrumentState(key, instruments[key])
        for key in models
    }

    log.info(f"\n{'═'*55}")
    log.info(f"  XGBoost Live Trader STARTED (FIX)")
    log.info(f"  Instruments : {list(states.keys())}")
    log.info(f"  Risk/trade  : {RISK_PER_TRADE:.1%}")
    log.info(f"  Confidence  : {MIN_CONFIDENCE}")
    log.info(f"  Poll every  : {POLL_INTERVAL_SECONDS}s")
    log.info(f"{'═'*55}\n")

    # ── Main loop ─────────────────────────────────────────────────────
    # Connectivity state: silently retry up to 5 times before declaring
    # offline. Once offline, check every 60s — no repeated log spam.
    _offline_fail_count = 0   # consecutive failed connectivity checks
    _offline = False           # True once confirmed offline (5 failures)
    _last_firestore_heartbeat = 0.0
    FIRESTORE_HEARTBEAT_SEC = int(os.environ.get("FIREBASE_HEARTBEAT_SEC", "300"))

    def _push_firestore_heartbeat(is_offline: bool = False) -> None:
        nonlocal _last_firestore_heartbeat
        if _firestore_logger is None or not _firestore_logger.enabled:
            return
        now = time.time()
        if now - _last_firestore_heartbeat < FIRESTORE_HEARTBEAT_SEC:
            return
        _last_firestore_heartbeat = now
        try:
            from integrations.firestore_logger import build_account_summary

            broker.refresh_positions()
            positions = broker.get_bot_positions()
            summary = build_account_summary(
                broker_name=broker.name,
                broker_backend="fix",
                balance=broker.get_account_balance(),
                currency=broker.get_account_currency(),
                open_positions=len(positions),
                instruments=list(states.keys()),
                dry_run=args.dry_run,
                offline=is_offline,
            )
            _firestore_logger.update_account(summary)
        except Exception as exc:
            log.debug("Firestore heartbeat skipped: %s", exc)

    _push_firestore_heartbeat(is_offline=False)

    while True:
        try:
            # ── Connectivity guard ───────────────────────────────────
            if not broker.ensure_connected():
                _offline_fail_count += 1
                if _offline_fail_count < 5:
                    # Silent retry — not enough failures to declare offline
                    time.sleep(5)
                    continue
                # 5+ consecutive failures — genuinely offline
                if not _offline:
                    log.warning(
                        "No internet connection detected. "
                        "Bot is paused and will resume automatically."
                    )
                    _offline = True
                    _push_firestore_heartbeat(is_offline=True)
                time.sleep(60)          # check every minute, no logging
                _offline_fail_count = 0  # reset counter for next round
                continue

            # Connected — reset offline state
            _offline_fail_count = 0
            if _offline:
                log.info("Internet restored. Resuming trading.")
                _offline = False

            _push_firestore_heartbeat(is_offline=False)

            # ── Normal trading poll ──────────────────────────────────
            for key, state in states.items():
                run_instrument_loop(state, models[key], dry_run=args.dry_run)

        except KeyboardInterrupt:
            log.info("\nKeyboard interrupt — shutting down.")
            broker.shutdown()
            break
        except Exception as exc:
            log.error(f"Unexpected error: {exc}", exc_info=True)
            log.info("Sleeping 60s before retry...")
            time.sleep(60)


if __name__ == "__main__":
    main()